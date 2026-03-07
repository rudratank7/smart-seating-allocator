from django.shortcuts import render
from django.http import HttpResponse
import io
import re
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side as BorderSide
from openpyxl.cell.cell import MergedCell

def natural_sort_key(s):
    """
    Helper function to sort strings containing numbers naturally.
    e.g., 'B001', 'B002', 'B101', 'B201' instead of pure alphabetical.
    """
    return[int(text) if text.isdigit() else str(text).lower() for text in re.split(r'(\d+)', str(s))]

def allocate_students_smart_blocks(classrooms, students):
    """
    Strict 2-Block Priority Logic:
    1. Primarily splits room usable capacity evenly into just Block 1 and Block 2.
    2. Fills ALL Block 1s across the college first, then ALL Block 2s.
    3. If a branch leaves empty space in Block 1 or Block 2, that remaining space 
       is dynamically assigned as Block 3 or Block 4 respectively.
    4. Block 3 and Block 4 are only used in very rare overflow cases and natively 
       guarantee that only small fragments (< 49% capacity) will be put there.
    """
    students_remaining = students.copy() 
    branch_rooms = defaultdict(list)
    
    # Filter rooms > 2 capacity and strictly sort alphanumerically
    sorted_rooms = sorted(
        [r for r in classrooms if r[1] > 2], 
        key=lambda x: natural_sort_key(x[0])
    )
    
    room_data =[]
    for room_name, room_cap in sorted_rooms:
        usable_cap = max(0, room_cap - 2)
        
        # Room is primarily divided into only two blocks (50/50 split)
        b1_cap = usable_cap // 2 + usable_cap % 2
        b2_cap = usable_cap // 2
        
        room_data.append({
            'room_name': room_name,
            'room_cap': room_cap,
            'blocks':[
                {'cap': b1_cap, 'branch': None, 'count': 0}, # Block 1 (index 0)
                {'cap': b2_cap, 'branch': None, 'count': 0}, # Block 2 (index 1)
                {'cap': 0, 'branch': None, 'count': 0},      # Block 3 (index 2) - initially 0
                {'cap': 0, 'branch': None, 'count': 0}       # Block 4 (index 3) - initially 0
            ]
        })

    # Create a global sequence prioritizing ALL Block 1s first, then ALL Block 2s.
    main_slots_b1 =[(r_idx, 0) for r_idx in range(len(room_data))]
    main_slots_b2 = [(r_idx, 1) for r_idx in range(len(room_data))]
    main_slots = main_slots_b1 + main_slots_b2
    
    # Leftover queues (Very Rare and Very Very Rare)
    leftover_slots_b3 =[]
    leftover_slots_b4 =[]

    # Sort branches by strength descending (largest first)
    branches = sorted([(b, c) for b, c in students_remaining.items() if c > 0], 
        key=lambda x: x[1], 
        reverse=True
    )

    main_idx = 0
    b3_idx = 0
    b4_idx = 0

    # Process one branch at a time strictly
    for branch_name, count in branches:
        while count > 0:
            assigned = False
            
            # --- TIER 1: Use Main Blocks (Block 1 and Block 2) ---
            if main_idx < len(main_slots):
                r_idx, b_idx = main_slots[main_idx]
                main_idx += 1
                
                room = room_data[r_idx]
                cap = room['blocks'][b_idx]['cap']
                
                if cap <= 0:
                    continue
                    
                take = min(count, cap)
                
                # Assign students
                room['blocks'][b_idx]['branch'] = branch_name
                room['blocks'][b_idx]['count'] = take
                
                # Track for summary
                if room['room_name'] not in branch_rooms[branch_name]:
                    branch_rooms[branch_name].append(room['room_name'])

                count -= take
                students_remaining[branch_name] -= take
                
                # If block was not completely filled, generate a Leftover Block 
                # (This natively ensures it holds < 49% of room space)
                leftover = cap - take
                if leftover > 0:
                    if b_idx == 0:  # Leftover of Block 1 becomes Block 3
                        room['blocks'][2]['cap'] = leftover
                        leftover_slots_b3.append((r_idx, 2))
                    elif b_idx == 1:  # Leftover of Block 2 becomes Block 4
                        room['blocks'][3]['cap'] = leftover
                        leftover_slots_b4.append((r_idx, 3))
                        
                assigned = True

            # --- TIER 2: Main Blocks Exhausted. Use Block 3 (Rare) ---
            elif b3_idx < len(leftover_slots_b3):
                r_idx, b_idx = leftover_slots_b3[b3_idx]
                b3_idx += 1
                
                room = room_data[r_idx]
                cap = room['blocks'][b_idx]['cap']
                
                take = min(count, cap)
                room['blocks'][b_idx]['branch'] = branch_name
                room['blocks'][b_idx]['count'] = take
                
                if room['room_name'] not in branch_rooms[branch_name]:
                    branch_rooms[branch_name].append(room['room_name'])

                count -= take
                students_remaining[branch_name] -= take
                
                assigned = True

            # --- TIER 3: Block 3 Exhausted. Use Block 4 (Very Rare) ---
            elif b4_idx < len(leftover_slots_b4):
                r_idx, b_idx = leftover_slots_b4[b4_idx]
                b4_idx += 1
                
                room = room_data[r_idx]
                cap = room['blocks'][b_idx]['cap']
                
                take = min(count, cap)
                room['blocks'][b_idx]['branch'] = branch_name
                room['blocks'][b_idx]['count'] = take
                
                if room['room_name'] not in branch_rooms[branch_name]:
                    branch_rooms[branch_name].append(room['room_name'])

                count -= take
                students_remaining[branch_name] -= take
                
                assigned = True

            # If no slots are available anywhere at all, break to prevent infinite loop
            if not assigned:
                break

    # Format allocations for the Excel generator exactly as required
    allocations =[]
    for room in room_data:
        allocated_blocks =[]
        total_students_in_room = 0
        active_blocks_count = 0
        
        for b in room['blocks']:
            if b['branch'] and b['count'] > 0:
                allocated_blocks.append((b['branch'], b['count']))
                total_students_in_room += b['count']
                active_blocks_count += 1
            else:
                allocated_blocks.append(("", 0))

        if total_students_in_room > 0:
            allocations.append((
                room['room_name'],
                room['room_cap'],
                max(0, room['room_cap'] - total_students_in_room),
                allocated_blocks,
                total_students_in_room,
                active_blocks_count
            ))

    final_students_remaining = {b: c for b, c in students_remaining.items()}
    
    return allocations, branch_rooms, final_students_remaining


def generate_excel(allocations, branch_rooms, students_remaining, original_students, user_data=""):
    """
    Generates the Excel file with structured groupings and minimal fragmentation.
    user_data: String, numbers, or any data entered by user to be placed in row 3
    """
    wb = Workbook()
    ws = wb.active
    
    # --- STYLES ---
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin_border = Border(
        left=BorderSide(style="thin"), right=BorderSide(style="thin"),
        top=BorderSide(style="thin"), bottom=BorderSide(style="thin")
    )

    # --- TOP HEADERS ---
    ws.append(["P P SAVANI UNIVERSITY"])
    ws.append(["School of Engineering"])
    ws.append([user_data]) 
    ws.append(["Seating Arrangement"])
   
    max_col_letter = 'N'
    for r in range(1, 5):
        ws.merge_cells(f'A{r}:{max_col_letter}{r}')
        cell = ws.cell(row=r, column=1)
        cell.alignment = center_align
        cell.font = Font(bold=True, size=14) if r in[1, 2, 4] else Font(bold=True)
    
    ws.append([]) 
    
    # --- TABLE HEADERS ---
    left_headers =[
        "Sr No", "Class Room", "Class Capacity", "Available",
        "Block 1", "Count", "Block 2", "Count", "Block 3", "Count", "Block 4", "Count",
        "Total students", "Total Block"
    ]
    ws.append(left_headers)
    
    # --- FILL ALLOCATION DATA ---
    current_row = 7
    for sr_no, (room, capacity, available, allocated, total_students, total_blocks) in enumerate(allocations, start=1):
        row_data = [sr_no, room, capacity, available]
        
        # 'allocated' possesses precisely 4 elements ensuring B1, B2, B3, B4 alignment
        for branch_name, count in allocated:
            if count > 0:
                row_data.extend([branch_name, count])
            else:
                row_data.extend(["", ""])
        
        row_data.extend([total_students, total_blocks])
        ws.append(row_data)
        current_row += 1

    # --- TOTALS ROW ---
    ws.append([''] * len(left_headers))
    current_row += 1
    
    totals_row = [''] * len(left_headers)
    totals_row[11] = "Totals" 
    totals_row[12] = f'=SUM(M7:M{current_row - 2})'
    totals_row[13] = f'=SUM(N7:N{current_row - 2})'

    ws.append(totals_row)
    current_row += 1
    
    # --- SUMMARY TABLE (RIGHT SIDE) ---
    start_col = 16 
    
    right_headers =["Branch", "Original Students", "Remaining Students", "Class Rooms Allocated"]

    for j, h in enumerate(right_headers):
        cell = ws.cell(row=6, column=start_col + j, value=h)
        cell.alignment = center_align
        cell.border = thin_border
        cell.font = bold_font
        cell.fill = yellow_fill 
        
    row_idx = 7
    
    # Preserve input order from original file for predictability
    for branch, original_count in original_students.items():
        if original_count > 0:
            remaining_count = students_remaining.get(branch, 0)
            
            # Use natural sort key again here to ensure room lists read cleanly like "A1, A2, B1"
            allocated_rooms = sorted(branch_rooms.get(branch,[]), key=natural_sort_key)
            
            ws.cell(row=row_idx, column=start_col, value=branch)
            ws.cell(row=row_idx, column=start_col + 1, value=original_count)
            ws.cell(row=row_idx, column=start_col + 2, value=remaining_count)
            ws.cell(row=row_idx, column=start_col + 3, value=", ".join(allocated_rooms))
            row_idx += 1

    # --- FORMATTING ---
    for row in ws.iter_rows(min_row=6, max_row=current_row, min_col=1, max_col=14):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            if cell.row == 6:
                cell.font = bold_font
                cell.fill = header_fill
            elif cell.row == current_row - 1: 
                 cell.font = bold_font

    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=16, max_col=19):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
                
    for col in ws.columns:
        max_length = 0
        try:
            col_letter = col[0].column_letter
        except AttributeError:
            continue

        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            try:
                if cell.value:
                    val_str = str(cell.value)
                    if val_str.startswith('='):
                        max_length = max(max_length, 10)
                    elif ',' in val_str:
                        max_length = max(max_length, 30)
                    else:
                        max_length = max(max_length, len(val_str))
            except:
                pass
        
        if col_letter == 'A':
             ws.column_dimensions[col_letter].width = 10
        elif col_letter == 'B':
             ws.column_dimensions[col_letter].width = 20
        elif max_length > 0:
             ws.column_dimensions[col_letter].width = min(max_length + 3, 40)

    return wb


def upload_and_process(request):
    """
    Django view function handling the request logic.
    """
    if request.method == "GET":
        return render(request, "seating/upload.html")

    if 'class_file' not in request.FILES or 'student_file' not in request.FILES:
         return HttpResponse("Error: Both class and student files must be uploaded.", status=400)
        
    try:
        class_file = request.FILES["class_file"]
        student_file = request.FILES["student_file"]
        user_data = request.POST.get("user_data", "").strip()

        class_df = pd.read_excel(io.BytesIO(class_file.read()), engine='openpyxl', header=None)
        student_df = pd.read_excel(io.BytesIO(student_file.read()), engine='openpyxl', header=None)

        class_df = class_df.dropna(how='all')
        student_df = student_df.dropna(how='all')
        
        class_df = class_df.iloc[1:].copy()
        student_df = student_df.iloc[1:].copy()

        class_df.columns =["Class Room", "Class Capacity"]
        student_df.columns = ["Branch", "No. of Students"]

        class_df["Class Capacity"] = pd.to_numeric(class_df["Class Capacity"], errors='coerce').fillna(0).astype(int)
        student_df["No. of Students"] = pd.to_numeric(student_df["No. of Students"], errors='coerce').fillna(0).astype(int)

        classrooms = [(str(row["Class Room"]), row["Class Capacity"]) 
                      for _, row in class_df.iterrows() if pd.notna(row["Class Room"]) and row["Class Capacity"] > 0]
        
        students = {}
        for _, row in student_df.iterrows():
            if pd.notna(row["Branch"]) and row["No. of Students"] > 0:
                students[str(row["Branch"]).strip()] = row["No. of Students"]
        
        original_students = students.copy() 

        allocations, branch_rooms, students_remaining = allocate_students_smart_blocks(classrooms, students)
        
        wb = generate_excel(allocations, branch_rooms, students_remaining, original_students, user_data)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="output_data_final_seating.xlsx"'
        return response

    except Exception as e:
        import traceback
        error_message = f"An unexpected error occurred: {e}\n{traceback.format_exc()}"
        return HttpResponse(error_message, status=500)